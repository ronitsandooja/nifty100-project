import sqlite3
import yaml
import pandas as pd

from src.analytics.cagr import cagr, shift_year


def load_config(path="config/screener_config.yaml"):
    with open(path) as f:
        return yaml.safe_load(f)


def build_universe(db_path="db/nifty100.db"):
    conn = sqlite3.connect(db_path)

    fr = pd.read_sql("select * from financial_ratios", conn)
    latest_fr = fr.sort_values("year").groupby("company_id").tail(1)

    mc = pd.read_sql("select * from market_cap", conn)
    latest_mc = mc.sort_values("year").groupby("company_id").tail(1)

    pnl = pd.read_sql("select company_id, year, sales, net_profit from profitandloss", conn)
    latest_pnl = pnl.sort_values("year").groupby("company_id").tail(1)

    sectors = pd.read_sql("select company_id, broad_sector from sectors", conn)
    companies = pd.read_sql("select id as company_id, company_name from companies", conn)

    conn.close()

    universe = latest_fr.merge(
        latest_mc[["company_id", "market_cap_crore", "pe_ratio", "pb_ratio", "dividend_yield_pct"]],
        on="company_id", how="left"
    )
    universe = universe.merge(latest_pnl[["company_id", "sales", "net_profit"]], on="company_id", how="left")
    universe = universe.merge(sectors, on="company_id", how="left")
    universe = universe.merge(companies, on="company_id", how="left")

    return universe


def passes_filter(row, metric_key, threshold, config):
    metric = config["metrics"][metric_key]
    column = metric["column"]
    direction = metric["direction"]

    if metric.get("skip_financials") and row.get("broad_sector") == "Financials":
        return True

    value = row.get(column)

    if metric.get("debt_free_passes") and pd.isna(value):
        return True

    if pd.isna(value):
        return False

    if direction == "min":
        return value >= threshold
    if direction == "max":
        return value <= threshold
    if direction == "exact":
        return value == threshold

    return False


def apply_filters(universe, filters, config):
    result = universe.copy()
    for metric_key, threshold in filters.items():
        mask = result.apply(lambda row: passes_filter(row, metric_key, threshold, config), axis=1)
        result = result[mask]
    return result


def turnaround_watch(universe, pnl_history, fr_history):
    matches = []

    for _, row in universe.iterrows():
        company_id = row["company_id"]

        if pd.isna(row.get("free_cash_flow_cr")) or row["free_cash_flow_cr"] <= 0:
            continue

        company_pnl = pnl_history[pnl_history["company_id"] == company_id].sort_values("year")
        if company_pnl.empty:
            continue

        latest_year = company_pnl["year"].iloc[-1]
        start_year = shift_year(latest_year, 3)
        start_row = company_pnl[company_pnl["year"] == start_year]
        if start_row.empty:
            continue

        latest_sales = company_pnl["sales"].iloc[-1]
        rev_cagr, flag = cagr(start_row["sales"].iloc[0], latest_sales, 3)
        if rev_cagr is None or rev_cagr <= 10:
            continue

        company_fr = fr_history[fr_history["company_id"] == company_id].sort_values("year")
        if len(company_fr) < 2:
            continue

        latest_de = company_fr["debt_to_equity"].iloc[-1]
        prior_de = company_fr["debt_to_equity"].iloc[-2]
        if pd.isna(latest_de) or pd.isna(prior_de) or latest_de >= prior_de:
            continue

        matches.append(row)

    return pd.DataFrame(matches)


def run_preset(universe, preset_key, config, pnl_history=None, fr_history=None):
    preset = config["presets"][preset_key]

    if preset.get("special"):
        return turnaround_watch(universe, pnl_history, fr_history)

    return apply_filters(universe, preset["filters"], config)


def scale_to_100(value, low, high):
    if value is None or pd.isna(value) or low is None or high is None or pd.isna(low) or pd.isna(high) or high == low:
        return None
    scaled = (value - low) / (high - low) * 100
    return max(0, min(100, scaled))


def weighted_average(parts):
    available = [(score, weight) for score, weight in parts if score is not None]
    if not available:
        return None
    total_weight = sum(weight for _, weight in available)
    return sum(score * weight for score, weight in available) / total_weight


def compute_fcf_cagr(company_fr):
    if len(company_fr) < 2:
        return None
    latest_year = company_fr["year"].iloc[-1]
    start_year = shift_year(latest_year, 5)
    start_row = company_fr[company_fr["year"] == start_year]
    if start_row.empty:
        return None
    latest_fcf = company_fr["free_cash_flow_cr"].iloc[-1]
    value, flag = cagr(start_row["free_cash_flow_cr"].iloc[0], latest_fcf, 5)
    return value


def add_helper_columns(universe, fr_history):
    universe = universe.copy()

    fcf_cagr_values = []
    cfo_pat_values = []

    for _, row in universe.iterrows():
        company_fr = fr_history[fr_history["company_id"] == row["company_id"]].sort_values("year")
        fcf_cagr_values.append(compute_fcf_cagr(company_fr))

        cfo = row.get("cash_from_operations_cr")
        pat = row.get("net_profit")
        if pat not in (None, 0) and not pd.isna(pat) and cfo is not None and not pd.isna(cfo):
            cfo_pat_values.append(cfo / pat)
        else:
            cfo_pat_values.append(None)

    universe["_fcf_cagr"] = fcf_cagr_values
    universe["_cfo_pat"] = cfo_pat_values
    return universe


def compute_composite_scores(universe, fr_history, group_by_sector=False):
    universe = add_helper_columns(universe, fr_history)

    groups = universe.groupby("broad_sector") if group_by_sector else [(None, universe)]

    all_scores = pd.Series(index=universe.index, dtype="float64")

    bound_columns = [
        "return_on_equity_pct", "return_on_capital_employed_pct", "net_profit_margin_pct",
        "revenue_cagr_5yr", "pat_cagr_5yr", "_fcf_cagr", "_cfo_pat"
    ]

    for _, group in groups:
        bounds = {col: group[col].quantile([0.1, 0.9]) for col in bound_columns}
        de_inverted = group["debt_to_equity"].apply(lambda x: -x if pd.notna(x) else None)
        bounds["de_inverted"] = de_inverted.quantile([0.1, 0.9])
        icr_filled = group["interest_coverage"].fillna(group["interest_coverage"].max())
        bounds["icr"] = icr_filled.quantile([0.1, 0.9])

        for idx, row in group.iterrows():
            roe_score = scale_to_100(row["return_on_equity_pct"], bounds["return_on_equity_pct"].iloc[0], bounds["return_on_equity_pct"].iloc[1])
            roce_score = scale_to_100(row["return_on_capital_employed_pct"], bounds["return_on_capital_employed_pct"].iloc[0], bounds["return_on_capital_employed_pct"].iloc[1])
            npm_score = scale_to_100(row["net_profit_margin_pct"], bounds["net_profit_margin_pct"].iloc[0], bounds["net_profit_margin_pct"].iloc[1])
            profitability = weighted_average([(roe_score, 15), (roce_score, 10), (npm_score, 10)])

            fcf_cagr_score = scale_to_100(row["_fcf_cagr"], bounds["_fcf_cagr"].iloc[0], bounds["_fcf_cagr"].iloc[1])
            cfo_pat_score = scale_to_100(row["_cfo_pat"], bounds["_cfo_pat"].iloc[0], bounds["_cfo_pat"].iloc[1])
            fcf_positive_score = 100 if (row.get("free_cash_flow_cr") or 0) > 0 else 0
            cash_quality = weighted_average([(fcf_cagr_score, 15), (cfo_pat_score, 10), (fcf_positive_score, 5)])

            rev_cagr_score = scale_to_100(row["revenue_cagr_5yr"], bounds["revenue_cagr_5yr"].iloc[0], bounds["revenue_cagr_5yr"].iloc[1])
            pat_cagr_score = scale_to_100(row["pat_cagr_5yr"], bounds["pat_cagr_5yr"].iloc[0], bounds["pat_cagr_5yr"].iloc[1])
            growth = weighted_average([(rev_cagr_score, 10), (pat_cagr_score, 10)])

            de_value = -row["debt_to_equity"] if pd.notna(row["debt_to_equity"]) else None
            de_score = scale_to_100(de_value, bounds["de_inverted"].iloc[0], bounds["de_inverted"].iloc[1])
            icr_value = row["interest_coverage"] if pd.notna(row["interest_coverage"]) else group["interest_coverage"].max()
            icr_score = scale_to_100(icr_value, bounds["icr"].iloc[0], bounds["icr"].iloc[1])
            leverage = weighted_average([(de_score, 10), (icr_score, 5)])

            total = weighted_average([(profitability, 35), (cash_quality, 30), (growth, 20), (leverage, 15)])
            all_scores[idx] = round(total, 1) if total is not None else None

    universe = universe.drop(columns=["_fcf_cagr", "_cfo_pat"])
    universe["composite_score"] = all_scores
    return universe


DISPLAY_COLUMNS = [
    "company_id", "company_name", "broad_sector",
    "return_on_equity_pct", "return_on_capital_employed_pct", "net_profit_margin_pct",
    "operating_profit_margin_pct", "debt_to_equity", "interest_coverage", "asset_turnover",
    "free_cash_flow_cr", "earnings_per_share", "book_value_per_share", "dividend_payout_ratio_pct",
    "revenue_cagr_5yr", "pat_cagr_5yr", "eps_cagr_5yr",
    "pe_ratio", "pb_ratio", "dividend_yield_pct", "market_cap_crore", "composite_score"
]


def export_screener_output(universe, config, pnl_history, fr_history, path="output/screener_output.xlsx"):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    wb = Workbook()
    wb.remove(wb.active)

    for preset_key, preset in config["presets"].items():
        result = run_preset(universe, preset_key, config, pnl_history, fr_history)
        result = result.sort_values("composite_score", ascending=False)

        sheet_name = preset["name"][:31]
        ws = wb.create_sheet(sheet_name)

        ws.append(DISPLAY_COLUMNS)

        filter_columns = set()
        if not preset.get("special"):
            for metric_key in preset["filters"]:
                filter_columns.add(config["metrics"][metric_key]["column"])

        for _, row in result.iterrows():
            values = [row.get(col) for col in DISPLAY_COLUMNS]
            ws.append(values)

            excel_row = ws.max_row
            for col_idx, col_name in enumerate(DISPLAY_COLUMNS, start=1):
                if col_name in filter_columns:
                    ws.cell(row=excel_row, column=col_idx).fill = green_fill

    wb.save(path)


if __name__ == "__main__":
    import sqlite3

    config = load_config()
    universe = build_universe()

    conn = sqlite3.connect("db/nifty100.db")
    pnl_history = pd.read_sql("select company_id, year, sales, net_profit from profitandloss", conn)
    fr_history = pd.read_sql("select * from financial_ratios", conn)
    conn.close()

    universe = compute_composite_scores(universe, fr_history)
    export_screener_output(universe, config, pnl_history, fr_history)

    print("output/screener_output.xlsx written")
    for preset_key in config["presets"]:
        result = run_preset(universe, preset_key, config, pnl_history, fr_history)
        print(f"{preset_key}: {len(result)} companies")
