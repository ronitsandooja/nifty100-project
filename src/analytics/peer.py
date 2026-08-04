import sqlite3
import pandas as pd

METRICS = {
    "roe": "return_on_equity_pct",
    "roce": "return_on_capital_employed_pct",
    "npm": "net_profit_margin_pct",
    "de": "debt_to_equity",
    "fcf": "free_cash_flow_cr",
    "pat_cagr_5yr": "pat_cagr_5yr",
    "revenue_cagr_5yr": "revenue_cagr_5yr",
    "eps_cagr_5yr": "eps_cagr_5yr",
    "interest_coverage": "interest_coverage",
    "asset_turnover": "asset_turnover",
}

INVERTED_METRICS = {"de"}


def load_peer_groups(path="data/supporting/peer_groups.xlsx"):
    return pd.read_excel(path, header=0)


def latest_ratios(db_path="db/nifty100.db"):
    conn = sqlite3.connect(db_path)
    fr = pd.read_sql("select * from financial_ratios", conn)
    conn.close()
    return fr.sort_values("year").groupby("company_id").tail(1)


def percent_rank(series):
    n = len(series)
    if n <= 1:
        return pd.Series([1.0] * n, index=series.index)
    ranks = series.rank(method="min", ascending=True)
    return (ranks - 1) / (n - 1)


def compute_percentiles(peer_groups_df, ratios_df):
    rows = []

    for group_name, group in peer_groups_df.groupby("peer_group_name"):
        member_ids = group["company_id"].tolist()
        group_ratios = ratios_df[ratios_df["company_id"].isin(member_ids)]

        for metric_key, column in METRICS.items():
            values = group_ratios[["company_id", "year", column]].dropna(subset=[column])

            if values.empty:
                continue

            rank = percent_rank(values[column])

            if metric_key in INVERTED_METRICS:
                percentile = 1 - rank
            else:
                percentile = rank

            values = values.assign(percentile_rank=percentile)

            for _, row in values.iterrows():
                rows.append({
                    "company_id": row["company_id"],
                    "peer_group_name": group_name,
                    "metric": metric_key,
                    "value": row[column],
                    "percentile_rank": round(row["percentile_rank"], 4),
                    "year": row["year"],
                })

    return pd.DataFrame(rows)


def companies_without_peer_group(companies_df, peer_groups_df):
    grouped_ids = set(peer_groups_df["company_id"])
    missing = companies_df[~companies_df["id"].isin(grouped_ids)]
    return missing["id"].tolist()


def peer_message(company_id, peer_groups_df):
    if company_id not in set(peer_groups_df["company_id"]):
        return "No peer group assigned"
    return None


def save_peer_percentiles(percentiles_df, db_path="db/nifty100.db"):
    conn = sqlite3.connect(db_path)
    conn.execute("delete from peer_percentiles")
    percentiles_df.to_sql("peer_percentiles", conn, if_exists="append", index=False)
    conn.commit()
    conn.close()


def export_peer_comparison(peer_groups_df, companies_df, percentiles_df, path="output/peer_comparison.xlsx"):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gold_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

    wb = Workbook()
    wb.remove(wb.active)

    for group_name, group in peer_groups_df.groupby("peer_group_name"):
        ws = wb.create_sheet(group_name[:31])

        header = ["company_id", "company_name"]
        for metric_key in METRICS:
            header.append(metric_key)
            header.append(f"{metric_key}_percentile")
        ws.append(header)

        member_ids = group["company_id"].tolist()
        benchmark_ids = set(group[group["is_benchmark"] == 1]["company_id"])
        values_by_metric = {m: [] for m in METRICS}

        for company_id in member_ids:
            company_row = companies_df[companies_df["id"] == company_id]
            company_name = company_row["company_name"].iloc[0] if not company_row.empty else ""

            row_values = [company_id, company_name]
            for metric_key in METRICS:
                p = percentiles_df[
                    (percentiles_df["company_id"] == company_id) &
                    (percentiles_df["peer_group_name"] == group_name) &
                    (percentiles_df["metric"] == metric_key)
                ]
                value = p["value"].iloc[0] if not p.empty else None
                percentile = p["percentile_rank"].iloc[0] if not p.empty else None

                row_values.append(value)
                row_values.append(percentile)

                if value is not None:
                    values_by_metric[metric_key].append(value)

            ws.append(row_values)
            excel_row = ws.max_row

            if company_id in benchmark_ids:
                for col_idx in range(1, len(header) + 1):
                    ws.cell(row=excel_row, column=col_idx).fill = gold_fill
            else:
                for metric_idx in range(len(METRICS)):
                    pct_col = 3 + metric_idx * 2 + 1
                    cell = ws.cell(row=excel_row, column=pct_col)
                    if cell.value is not None:
                        if cell.value >= 0.75:
                            cell.fill = green_fill
                        elif cell.value <= 0.25:
                            cell.fill = red_fill
                        else:
                            cell.fill = yellow_fill

        summary_row = ["MEDIAN", ""]
        for metric_key in METRICS:
            values = values_by_metric[metric_key]
            median = pd.Series(values).median() if values else None
            summary_row.append(median)
            summary_row.append(None)
        ws.append(summary_row)

    wb.save(path)


if __name__ == "__main__":
    conn = sqlite3.connect("db/nifty100.db")
    peer_groups_df = pd.read_sql("select * from peer_groups", conn)
    companies_df = pd.read_sql("select * from companies", conn)
    conn.close()

    ratios_df = latest_ratios()

    percentiles_df = compute_percentiles(peer_groups_df, ratios_df)
    save_peer_percentiles(percentiles_df)

    export_peer_comparison(peer_groups_df, companies_df, percentiles_df)

    missing = companies_without_peer_group(companies_df, peer_groups_df)

    print(f"peer_percentiles -> {len(percentiles_df)} rows")
    print(f"{peer_groups_df['peer_group_name'].nunique()} peer groups covered")
    print(f"{len(missing)} companies have no peer group assigned")
    print("output/peer_comparison.xlsx written")
